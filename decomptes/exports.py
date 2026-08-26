import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

LIBELLES_TYPE = {"garde": "Garde", "permanence": "Permanence", "astreinte": "Astreinte"}

HEURES_PAR_UNITE = 12
TAUX_UNITE_DH = {"medecin": 296, "infirmier": 140}
TAUX_PERMANENCE_DH = 5

ENTETE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ENTETE_FONT = Font(color="FFFFFF", bold=True)


def _unites_et_montant(decompte):
    """Unités de garde (1 unité = 12h) et montant en DH, uniquement pour
    médecins et infirmiers. Retourne (None, None) pour tout autre profil."""
    taux = TAUX_UNITE_DH.get(decompte.agent.fonction)
    if not taux:
        return None, None
    unites = decompte.total_heures / HEURES_PAR_UNITE
    montant = round(unites * taux, 2)
    return round(unites, 2), montant


def exporter_decomptes_excel(service, annee, mois, decomptes, type_activite="garde"):
    decomptes = list(decomptes)
    est_permanence = type_activite == "permanence"
    est_astreinte = type_activite == "astreinte"
    afficher_ramadan = (not est_permanence and not est_astreinte) and any(d.heures_ramadan for d in decomptes)
    libelle_type = LIBELLES_TYPE.get(type_activite, type_activite)

    wb = Workbook()
    ws = wb.active
    ws.title = "Décompte"

    if est_astreinte:
        entetes = ["Matricule", "Nom", "Prénom", "Jours ouvrables", "Jours fériés", "Unités"]
    elif est_permanence:
        entetes = ["Matricule", "Nom", "Prénom", "Ouvrable", "Férié", "Total", "Montant (DH)"]
    else:
        entetes = ["Matricule", "Nom", "Prénom", "Ouvrable", "Vendredi"]
        if afficher_ramadan:
            entetes.append("Ramadan")
        entetes += ["Week-end/férié", "Nuit", "Total", "Unités", "Montant (DH)"]

    ws.merge_cells(f"A1:{get_column_letter(len(entetes))}1")
    ws["A1"] = f"Décompte de {libelle_type.lower()} — {service.nom} — {MOIS_FR[mois]} {annee}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(entetes)
    for col_idx, _ in enumerate(entetes, start=1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = ENTETE_FILL
        c.font = ENTETE_FONT
        c.alignment = Alignment(horizontal="center")

    total = 0.0
    total_montant = 0.0
    for d in decomptes:
        if est_astreinte:
            ligne = [d.agent.matricule, d.agent.nom, d.agent.prenom, int(d.heures_ouvrable), int(d.heures_weekend_ferie), d.total_heures]
        elif est_permanence:
            montant = round(d.total_heures * TAUX_PERMANENCE_DH, 2)
            total_montant += montant
            ligne = [d.agent.matricule, d.agent.nom, d.agent.prenom, d.heures_ouvrable, d.heures_weekend_ferie, d.total_heures, montant]
        else:
            unites, montant = _unites_et_montant(d)
            ligne = [d.agent.matricule, d.agent.nom, d.agent.prenom, d.heures_ouvrable, d.heures_vendredi]
            if afficher_ramadan:
                ligne.append(d.heures_ramadan or None)
            ligne += [d.heures_weekend_ferie, d.heures_nuit, d.total_heures, unites, montant]
            if montant:
                total_montant += montant
        ws.append(ligne)
        total += d.total_heures

    ligne_num = ws.max_row + 1
    ws.cell(row=ligne_num, column=3, value="TOTAL").font = Font(bold=True)
    if est_astreinte:
        ws.cell(row=ligne_num, column=len(entetes), value=round(total, 2)).font = Font(bold=True)
    elif est_permanence:
        ws.cell(row=ligne_num, column=len(entetes) - 1, value=round(total, 2)).font = Font(bold=True)
        ws.cell(row=ligne_num, column=len(entetes), value=round(total_montant, 2)).font = Font(bold=True)
    else:
        ws.cell(row=ligne_num, column=len(entetes) - 2, value=total).font = Font(bold=True)
        ws.cell(row=ligne_num, column=len(entetes) - 1, value=round(total / HEURES_PAR_UNITE, 2)).font = Font(bold=True)
        if total_montant:
            ws.cell(row=ligne_num, column=len(entetes), value=round(total_montant, 2)).font = Font(bold=True)

    if est_astreinte:
        largeurs = [14, 16, 16, 16, 14, 10]
    elif est_permanence:
        largeurs = [14, 16, 16, 12, 12, 12, 14]
    else:
        largeurs = [14, 16, 16, 12, 12] + ([12] if afficher_ramadan else []) + [16, 10, 12, 10, 14]
    for idx, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largeur

    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return tampon


def exporter_decomptes_pdf(service, annee, mois, decomptes, type_activite="garde"):
    decomptes = list(decomptes)
    est_permanence = type_activite == "permanence"
    est_astreinte = type_activite == "astreinte"
    afficher_ramadan = (not est_permanence and not est_astreinte) and any(d.heures_ramadan for d in decomptes)
    libelle_type = LIBELLES_TYPE.get(type_activite, type_activite)

    tampon = io.BytesIO()
    doc = SimpleDocTemplate(tampon, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle("Titre", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=14)
    style_sous = ParagraphStyle("Sous", parent=styles["Normal"], alignment=TA_CENTER, fontSize=10)

    elements = [
        Paragraph("CHR de Fès — Hôpital Al Ghassani", style_sous),
        Paragraph(f"Bordereau de {libelle_type.lower()} — {service.nom}<br/>{MOIS_FR[mois]} {annee}", style_titre),
        Spacer(1, 0.8*cm),
    ]

    if est_astreinte:
        entetes = ["Matricule", "Nom et prénom", "Jours ouvr.", "Jours fériés", "Unités"]
    elif est_permanence:
        entetes = ["Matricule", "Nom et prénom", "Ouvrable", "Férié", "Total", "Montant (DH)"]
    else:
        entetes = ["Matricule", "Nom et prénom", "Ouvrable", "Vendredi"]
        if afficher_ramadan:
            entetes.append("Ramadan")
        entetes += ["W-E/férié", "Nuit", "Total", "Unités", "Montant (DH)"]

    donnees = [entetes]
    total = 0.0
    total_montant = 0.0
    for d in decomptes:
        if est_astreinte:
            ligne = [d.agent.matricule, d.agent.nom_complet, f"{int(d.heures_ouvrable)}", f"{int(d.heures_weekend_ferie)}", f"{d.total_heures:g}"]
        elif est_permanence:
            montant = round(d.total_heures * TAUX_PERMANENCE_DH, 2)
            total_montant += montant
            ligne = [d.agent.matricule, d.agent.nom_complet, f"{d.heures_ouvrable:g}", f"{d.heures_weekend_ferie:g}", f"{d.total_heures:g}", f"{montant:g}"]
        else:
            unites, montant = _unites_et_montant(d)
            ligne = [d.agent.matricule, d.agent.nom_complet, f"{d.heures_ouvrable:g}", f"{d.heures_vendredi:g}"]
            if afficher_ramadan:
                ligne.append(f"{d.heures_ramadan:g}" if d.heures_ramadan else "")
            ligne += [
                f"{d.heures_weekend_ferie:g}", f"{d.heures_nuit:g}", f"{d.total_heures:g}",
                f"{unites:g}" if unites is not None else "—",
                f"{montant:g}" if montant is not None else "—",
            ]
            if montant:
                total_montant += montant
        donnees.append(ligne)
        total += d.total_heures

    if est_astreinte:
        ligne_total = ["", "TOTAL"] + [""] * (len(entetes) - 3) + [f"{total:g}"]
    elif est_permanence:
        ligne_total = ["", "TOTAL", "", "", f"{total:g}", f"{total_montant:g}"]
    else:
        nb_blancs = len(entetes) - 5
        ligne_total = ["", "TOTAL"] + [""] * nb_blancs + [
            f"{total:g}", f"{total / HEURES_PAR_UNITE:g}", f"{total_montant:g}" if total_montant else "—",
        ]
    donnees.append(ligne_total)

    if est_astreinte:
        largeurs = [2.6*cm, 6*cm, 2.8*cm, 2.8*cm, 2.4*cm]
    elif est_permanence:
        largeurs = [2.4*cm, 5.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.5*cm]
    else:
        largeurs = [2.1*cm, 3.8*cm, 1.8*cm, 1.8*cm] + ([1.8*cm] if afficher_ramadan else []) + [1.9*cm, 1.6*cm, 1.8*cm, 1.6*cm, 2.2*cm]

    table = Table(donnees, repeatRows=1, colWidths=largeurs)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(table)
    doc.build(elements)
    tampon.seek(0)
    return tampon


def exporter_trimestre_excel(service, annee, trimestre, decomptes, type_activite):
    """Export Excel pour la vue trimestrielle en lecture seule (astreinte, ou
    permanence). `decomptes` est une liste de dicts (voir DecompteTrimestrielView),
    pas des instances de Decompte."""
    est_astreinte = type_activite == "astreinte"
    libelle_type = LIBELLES_TYPE.get(type_activite, type_activite)
    nom_trimestre = f"T{trimestre}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Décompte"

    entetes = ["Matricule", "Nom et prénom", "Jours ouvrables", "Jours fériés", "Unités"] if est_astreinte \
        else ["Matricule", "Nom et prénom", "Ouvrable", "Férié", "Total"]

    ws.merge_cells(f"A1:{get_column_letter(len(entetes))}1")
    ws["A1"] = f"Décompte de {libelle_type.lower()} — {service.nom} — {nom_trimestre} {annee}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(entetes)
    for col_idx, _ in enumerate(entetes, start=1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = ENTETE_FILL
        c.font = ENTETE_FONT
        c.alignment = Alignment(horizontal="center")

    total = 0.0
    for d in decomptes:
        if est_astreinte:
            ligne = [d['agent_matricule'], d['agent_nom'], int(d['heures_ouvrable']), int(d['heures_weekend_ferie']), d['total_heures']]
        else:
            ligne = [d['agent_matricule'], d['agent_nom'], d['heures_ouvrable'], d['heures_weekend_ferie'], d['total_heures']]
        ws.append(ligne)
        total += d['total_heures']

    ligne_num = ws.max_row + 1
    ws.cell(row=ligne_num, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=ligne_num, column=len(entetes), value=round(total, 2)).font = Font(bold=True)

    for idx, largeur in enumerate([14, 20, 16, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largeur

    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return tampon


def exporter_trimestre_pdf(service, annee, trimestre, decomptes, type_activite):
    est_astreinte = type_activite == "astreinte"
    libelle_type = LIBELLES_TYPE.get(type_activite, type_activite)
    nom_trimestre = f"T{trimestre}"

    tampon = io.BytesIO()
    doc = SimpleDocTemplate(tampon, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle("Titre", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=14)
    style_sous = ParagraphStyle("Sous", parent=styles["Normal"], alignment=TA_CENTER, fontSize=10)
    elements = [
        Paragraph("CHR de Fès — Hôpital Al Ghassani", style_sous),
        Paragraph(f"Bordereau de {libelle_type.lower()} — {service.nom}<br/>{nom_trimestre} {annee}", style_titre),
        Spacer(1, 0.8*cm),
    ]

    entetes = ["Matricule", "Nom et prénom", "Jours ouvr.", "Jours fériés", "Unités"] if est_astreinte \
        else ["Matricule", "Nom et prénom", "Ouvrable", "Férié", "Total"]

    donnees = [entetes]
    total = 0.0
    for d in decomptes:
        if est_astreinte:
            ligne = [d['agent_matricule'], d['agent_nom'], f"{int(d['heures_ouvrable'])}", f"{int(d['heures_weekend_ferie'])}", f"{d['total_heures']:g}"]
        else:
            ligne = [d['agent_matricule'], d['agent_nom'], f"{d['heures_ouvrable']:g}", f"{d['heures_weekend_ferie']:g}", f"{d['total_heures']:g}"]
        donnees.append(ligne)
        total += d['total_heures']
    donnees.append(["", "TOTAL", "", "", f"{total:g}"])

    table = Table(donnees, repeatRows=1, colWidths=[2.6*cm, 6*cm, 2.8*cm, 2.8*cm, 2.4*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(table)
    doc.build(elements)
    tampon.seek(0)
    return tampon
